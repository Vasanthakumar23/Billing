from __future__ import annotations

import re
import uuid
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi import File as FastAPIFile
from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.core.database import get_db
from app.models.student import Student
from app.models.student_balance_view import StudentBalanceView
from app.models.student_fee import StudentFee
from app.models.user import User
from app.models.enums import StudentStatus
from app.schemas.students import (
    StudentCreate,
    StudentListItem,
    StudentBalanceRead,
    StudentFeeRead,
    StudentFeeUpdate,
    StudentRead,
    StudentUpdate,
)


router = APIRouter()

_HEADER_RE = re.compile(r"[^a-z0-9]+")


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = _HEADER_RE.sub("", s)
    return s


def _find_col(headers: dict[str, int], keys: list[str]) -> int | None:
    for key in keys:
        idx = headers.get(key)
        if idx is not None:
            return idx
    return None


def _to_student_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip()
    if s == "":
        return None
    return Decimal(s)


@router.post("/import", response_model=dict)
async def import_students_from_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    file: UploadFile = FastAPIFile(...),
    mode: str = Query("upsert", pattern="^(upsert|create_only)$"),
    atomic: bool = Query(True),
) -> dict:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Only .xlsx files are supported")

    data = await file.read()
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Invalid Excel file")

    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Excel file is empty")

    headers: dict[str, int] = {}
    for idx, value in enumerate(header_row):
        key = _normalize_header(value)
        if key and key not in headers:
            headers[key] = idx

    student_code_col = _find_col(
        headers,
        [
            "studentcode",
            "studentid",
            "rollno",
            "rollnumber",
            "roll",
            "id",
        ],
    )
    name_col = _find_col(headers, ["name", "studentname"])
    class_col = _find_col(headers, ["classname", "class", "std", "standard"])
    section_col = _find_col(headers, ["section", "sec"])
    fee_col = _find_col(headers, ["expectedfeeamount", "expectedfee", "fee", "fees"])

    missing = []
    if student_code_col is None:
        missing.append("student_code/rollno")
    if name_col is None:
        missing.append("name")
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {', '.join(missing)}",
        )

    created = 0
    updated = 0
    fee_updated = 0
    errors: list[dict] = []

    # We process and write in one pass; if atomic=True we roll back on any error.
    try:
        for row_index, row in enumerate(rows, start=2):
            def cell(col: int | None) -> object:
                if col is None:
                    return None
                return row[col] if col < len(row) else None

            student_code = _to_student_code(cell(student_code_col))
            name_val = (cell(name_col) if name_col is not None else None)
            name = "" if name_val is None else str(name_val).strip()
            class_name = None
            if class_col is not None:
                v = cell(class_col)
                class_name = None if v is None or str(v).strip() == "" else str(v).strip()
            section = None
            if section_col is not None:
                v = cell(section_col)
                section = None if v is None or str(v).strip() == "" else str(v).strip()

            fee_cell = cell(fee_col) if fee_col is not None else None
            fee_cell_empty = fee_cell is None or str(fee_cell).strip() == ""
            if student_code == "" and name == "" and (class_name is None) and (section is None) and fee_cell_empty:
                continue  # empty-ish row

            if student_code == "":
                errors.append({"row": row_index, "error": "student_code/rollno is required"})
                continue
            if name == "":
                errors.append({"row": row_index, "error": "name is required"})
                continue

            try:
                fee = _to_decimal(cell(fee_col)) if fee_col is not None else None
            except (InvalidOperation, ValueError):
                errors.append({"row": row_index, "error": "fees must be a number"})
                continue
            if fee is None:
                fee = Decimal("0")
            if fee < 0:
                errors.append({"row": row_index, "error": "fees must be >= 0"})
                continue

            student = db.execute(select(Student).where(Student.student_code == student_code)).scalar_one_or_none()
            if student:
                if mode == "create_only":
                    errors.append({"row": row_index, "error": f"student_code '{student_code}' already exists"})
                    continue
                student.name = name
                student.class_name = class_name
                student.section = section
                updated += 1
            else:
                student = Student(
                    student_code=student_code,
                    name=name,
                    class_name=class_name,
                    section=section,
                )
                db.add(student)
                db.flush()
                created += 1

            fee_row = db.get(StudentFee, student.id)
            if not fee_row:
                fee_row = StudentFee(student_id=student.id)
                db.add(fee_row)
            fee_row.expected_fee_amount = fee
            fee_row.last_fee_updated_at = datetime.now(UTC)
            fee_row.last_fee_updated_by = current_user.id
            fee_updated += 1

        if errors and atomic:
            db.rollback()
            raise HTTPException(status_code=422, detail={"errors": errors})

        db.commit()
    except HTTPException:
        raise
    except Exception:
        db.rollback()
        raise

    return {
        "created": created,
        "updated": updated,
        "fee_updated": fee_updated,
        "errors": errors,
    }


@router.get("", response_model=dict)
def list_students(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
    search: str | None = None,
    status: StudentStatus | None = None,
    class_name: str | None = None,
    section: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> dict:
    stmt = select(Student)

    if search:
        s = f"%{search.lower()}%"
        stmt = stmt.where(
            or_(func.lower(Student.student_code).like(s), func.lower(Student.name).like(s))
        )
    if status is not None:
        stmt = stmt.where(Student.status == status)
    if class_name:
        stmt = stmt.where(Student.class_name == class_name)
    if section:
        stmt = stmt.where(Student.section == section)

    total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar_one()
    items = (
        db.execute(
            stmt.order_by(Student.student_code)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        .scalars()
        .all()
    )
    return {"items": [StudentRead.model_validate(s) for s in items], "total": total}


@router.get("/balances", response_model=dict)
def list_student_balances(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
    search: str | None = None,
    status: StudentStatus | None = None,
    class_name: str | None = None,
    section: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> dict:
    stmt = (
        select(
            Student.id,
            Student.student_code,
            Student.name,
            Student.class_name,
            Student.section,
            Student.status,
            StudentBalanceView.expected_fee,
            StudentBalanceView.paid_total,
            StudentBalanceView.pending,
        )
        .join(StudentBalanceView, StudentBalanceView.student_id == Student.id)
    )

    if search:
        s = f"%{search.lower()}%"
        stmt = stmt.where(
            or_(func.lower(Student.student_code).like(s), func.lower(Student.name).like(s))
        )
    if status is not None:
        stmt = stmt.where(Student.status == status)
    if class_name:
        stmt = stmt.where(Student.class_name == class_name)
    if section:
        stmt = stmt.where(Student.section == section)

    total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar_one()
    rows = (
        db.execute(
            stmt.order_by(Student.student_code)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        .all()
    )
    items = [
        StudentListItem(
            id=r.id,
            student_code=r.student_code,
            name=r.name,
            class_name=r.class_name,
            section=r.section,
            status=r.status,
            expected_fee=r.expected_fee,
            paid_total=r.paid_total,
            pending=r.pending,
        )
        for r in rows
    ]
    return {"items": items, "total": total}


@router.post("", response_model=StudentRead, status_code=201)
def create_student(
    payload: StudentCreate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentRead:
    existing = db.execute(select(Student).where(Student.student_code == payload.student_code)).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="student_code already exists")

    student = Student(
        student_code=payload.student_code,
        name=payload.name,
        class_name=payload.class_name,
        section=payload.section,
    )
    student.fee = StudentFee(expected_fee_amount=0)
    db.add(student)
    db.commit()
    db.refresh(student)
    return StudentRead.model_validate(student)


@router.get("/{student_id}", response_model=StudentRead)
def get_student(
    student_id: uuid.UUID,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentRead:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return StudentRead.model_validate(student)


@router.get("/{student_id}/balance", response_model=StudentBalanceRead)
def get_student_balance(
    student_id: uuid.UUID,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentBalanceRead:
    row = (
        db.execute(
            select(StudentBalanceView).where(StudentBalanceView.student_id == student_id)
        )
        .scalars()
        .one_or_none()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Student not found")
    return StudentBalanceRead.model_validate(row)


@router.patch("/{student_id}", response_model=StudentRead)
def update_student(
    student_id: uuid.UUID,
    payload: StudentUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentRead:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(student, key, value)

    db.commit()
    db.refresh(student)
    return StudentRead.model_validate(student)


@router.patch("/{student_id}/fee", response_model=StudentFeeRead)
def update_student_fee(
    student_id: uuid.UUID,
    payload: StudentFeeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StudentFeeRead:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    fee = db.get(StudentFee, student_id)
    if not fee:
        fee = StudentFee(student_id=student_id)
        db.add(fee)

    fee.expected_fee_amount = payload.expected_fee_amount
    fee.last_fee_updated_at = datetime.now(UTC)
    fee.last_fee_updated_by = current_user.id

    db.commit()
    db.refresh(fee)
    return StudentFeeRead.model_validate(fee)


@router.get("/{student_id}/fee", response_model=StudentFeeRead)
def get_student_fee(
    student_id: uuid.UUID,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StudentFeeRead:
    fee = db.get(StudentFee, student_id)
    if not fee:
        fee = StudentFee(student_id=student_id)
        db.add(fee)
        db.commit()
        db.refresh(fee)
    return StudentFeeRead.model_validate(fee)
